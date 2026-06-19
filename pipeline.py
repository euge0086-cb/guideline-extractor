"""
GUIDELINE REFERENCE PIPELINE
================================
Pipeline modular para extraer referencias de guías clínicas en PDF,
enriquecer con metadatos (PubMed + CrossRef) y exportar a Excel.

MÓDULOS:
  1. extract_references(pdf_path)  → lista de strings de referencias crudas
  2. enrich_reference(ref_text)    → dict con PMID, DOI, autores, año, etc.
  3. classify_reference(metadata)  → tipo: RCT_primario / RCT_secundario / meta-analisis / observacional / otro
  4. export_to_excel(records, out) → archivo .xlsx con hojas diferenciadas

USO:
  python guideline_pipeline.py <ruta_al_pdf> [output.xlsx]
"""

import re
import sys
import time
import json
import requests
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# MÓDULO 1: EXTRACCIÓN DE REFERENCIAS DEL PDF
# ─────────────────────────────────────────────

def extract_references_from_pdf(pdf_path: str) -> list[str]:
    """
    Extrae el bloque de referencias de un PDF de guía clínica.
    Soporta PDFs de una y dos columnas.
    Busca la sección 'References'/'Bibliografía' y extrae referencias numeradas.
    """
    all_lines = []
    in_refs = False

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_width = page.width

            # Intentar extracción por columnas (detecta layouts de 2 columnas)
            left_col = page.crop((0, 0, page_width / 2, page.height))
            right_col = page.crop((page_width / 2, 0, page_width, page.height))

            for col in [left_col, right_col]:
                words = col.extract_words()
                if not words:
                    continue
                # Agrupar palabras por línea (coordenada Y redondeada)
                line_map = {}
                for w in words:
                    y_key = round(w["top"])
                    line_map.setdefault(y_key, []).append(w["text"])

                for y_key in sorted(line_map.keys()):
                    line = " ".join(line_map[y_key]).strip()
                    if not line:
                        continue
                    # Detectar inicio de sección de referencias
                    if re.match(r'^(References|Bibliograf[ií]a|Referenci[ae]s)\s*$', line, re.IGNORECASE):
                        in_refs = True
                        continue
                    if in_refs:
                        all_lines.append(line)

    if not all_lines:
        # Fallback: extracción simple de texto completo
        print("[WARN] Fallback a extracción de texto simple.")
        with pdfplumber.open(pdf_path) as pdf:
            full_text = ""
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    full_text += t + "\n"
        for pat in [r'\bReferences\b', r'\bBibliograf[ií]a\b']:
            m = re.search(pat, full_text, re.IGNORECASE)
            if m:
                all_lines = full_text[m.end():].split("\n")
                break

    # Unir líneas en referencias completas (agrupando por número inicial)
    full_refs = []
    current = ""
    for line in all_lines:
        # Línea que empieza nueva referencia (número seguido de punto)
        if re.match(r'^\d{1,3}\.\s', line.strip()):
            if current:
                full_refs.append(re.sub(r'\s+', ' ', current).strip())
            current = line
        elif current:
            # Ignorar encabezados de página que se cuelen
            if re.match(r'^(Representativeness|C\.\s|European|804|805|802|803|800|801|799|798|797|796)', line):
                continue
            current += " " + line

    if current:
        full_refs.append(re.sub(r'\s+', ' ', current).strip())

    # Filtrar entradas vacías o demasiado cortas
    full_refs = [r for r in full_refs if len(r) > 25]

    print(f"[INFO] Referencias extraídas: {len(full_refs)}")
    return full_refs


# ─────────────────────────────────────────────
# MÓDULO 2: ENRIQUECIMIENTO VÍA APIs
# ─────────────────────────────────────────────

def search_pubmed(query: str) -> dict:
    """Busca en PubMed y devuelve PMID + metadatos básicos."""
    base = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/"
    
    # Buscar PMID
    search_url = f"{base}esearch.fcgi"
    params = {"db": "pubmed", "term": query, "retmax": 1, "retmode": "json"}
    try:
        r = requests.get(search_url, params=params, timeout=10)
        ids = r.json().get("esearchresult", {}).get("idlist", [])
        if not ids:
            return {}
        pmid = ids[0]
        
        # Obtener metadatos
        fetch_url = f"{base}efetch.fcgi"
        fetch_params = {"db": "pubmed", "id": pmid, "retmode": "xml", "rettype": "abstract"}
        rf = requests.get(fetch_url, params=fetch_params, timeout=10)
        xml = rf.text
        
        # Parsear campos clave del XML
        def extract_xml(tag, text):
            m = re.search(rf'<{tag}[^>]*>(.*?)</{tag}>', text, re.DOTALL)
            return m.group(1).strip() if m else ""
        
        title = extract_xml("ArticleTitle", xml)
        year = extract_xml("Year", xml) or extract_xml("MedlineDate", xml)[:4]
        journal = extract_xml("Title", xml)  # Journal Title
        
        # Autores
        authors_raw = re.findall(r'<LastName>(.*?)</LastName>.*?<ForeName>(.*?)</ForeName>', xml, re.DOTALL)
        authors = ", ".join([f"{ln} {fn[0]}." for ln, fn in authors_raw[:3]])
        if len(authors_raw) > 3:
            authors += " et al."
        
        # DOI desde PubMed
        doi_m = re.search(r'<ArticleId IdType="doi">(.*?)</ArticleId>', xml)
        doi = doi_m.group(1).strip() if doi_m else ""
        
        return {
            "pmid": pmid,
            "doi": doi,
            "title": title,
            "year": year[:4] if year else "",
            "journal": journal,
            "authors": authors,
            "source": "PubMed"
        }
    except Exception as e:
        return {}


def search_crossref(ref_text: str) -> dict:
    """Busca en CrossRef por texto de referencia libre."""
    url = "https://api.crossref.org/works"
    # Limpiar número de referencia
    clean = re.sub(r'^\d+[\.\s]+', '', ref_text).strip()[:200]
    params = {
        "query.bibliographic": clean,
        "rows": 1,
        "select": "DOI,title,author,published,container-title,type"
    }
    headers = {"User-Agent": "Guideline-Pipeline/1.0 (research tool)"}
    try:
        r = requests.get(url, params=params, headers=headers, timeout=10)
        items = r.json().get("message", {}).get("items", [])
        if not items:
            return {}
        item = items[0]
        
        doi = item.get("DOI", "")
        title = item.get("title", [""])[0] if item.get("title") else ""
        pub_type = item.get("type", "")
        
        authors_list = item.get("author", [])
        authors = ", ".join([
            f"{a.get('family', '')} {a.get('given', [''])[0]}." 
            if a.get('given') else a.get('family', '')
            for a in authors_list[:3]
        ])
        if len(authors_list) > 3:
            authors += " et al."
        
        year = ""
        pub = item.get("published", {}).get("date-parts", [[""]])
        if pub and pub[0]:
            year = str(pub[0][0])
        
        journal = ""
        ct = item.get("container-title", [])
        if ct:
            journal = ct[0]
        
        return {
            "doi": doi,
            "title": title,
            "year": year,
            "journal": journal,
            "authors": authors,
            "pub_type_raw": pub_type,
            "source": "CrossRef"
        }
    except Exception as e:
        return {}


def build_pubmed_query(ref_text: str) -> str:
    """Construye query PubMed desde texto de referencia."""
    # Extraer título tentativo: texto entre el primer punto y la revista
    clean = re.sub(r'^\d+[\.\s]+', '', ref_text).strip()
    # Intentar extraer primeras palabras significativas del título
    words = re.findall(r'\b[A-Za-z]{4,}\b', clean)
    query = " ".join(words[:8])
    # Añadir año si está presente
    year_m = re.search(r'\b(19|20)\d{2}\b', clean)
    if year_m:
        query += f"[Title/Abstract] AND {year_m.group()}[PDAT]"
    return query


def enrich_reference(ref_text: str, idx: int) -> dict:
    """
    Enriquece una referencia con metadatos desde PubMed y CrossRef.
    Combina los resultados priorizando PubMed para PMID y CrossRef para DOI.
    """
    record = {
        "ref_number": idx,
        "ref_raw": ref_text,
        "pmid": "",
        "doi": "",
        "title": "",
        "authors": "",
        "year": "",
        "journal": "",
        "study_type": "",
        "study_type_auto": "",
        "pubmed_url": "",
        "doi_url": "",
        "source_api": "",
        "notes": ""
    }

    time.sleep(0.35)  # Respetar rate limit NCBI (3 req/s sin API key)

    # 1. Intentar CrossRef primero (más tolerante a texto libre)
    cr = search_crossref(ref_text)
    if cr:
        record.update({k: v for k, v in cr.items() if v and k in record})
        record["source_api"] = "CrossRef"

    # 2. Buscar en PubMed para obtener PMID
    query = build_pubmed_query(ref_text)
    if query:
        time.sleep(0.35)
        pm = search_pubmed(query)
        if pm:
            # PubMed tiene prioridad para PMID y puede complementar datos
            if pm.get("pmid"):
                record["pmid"] = pm["pmid"]
                record["pubmed_url"] = f"https://pubmed.ncbi.nlm.nih.gov/{pm['pmid']}/"
            # Rellenar campos vacíos con datos de PubMed
            for k in ["doi", "title", "authors", "year", "journal"]:
                if not record[k] and pm.get(k):
                    record[k] = pm[k]
            record["source_api"] = "PubMed+CrossRef" if cr else "PubMed"

    # Construir URL de DOI
    if record["doi"]:
        record["doi_url"] = f"https://doi.org/{record['doi']}"

    return record


# ─────────────────────────────────────────────
# MÓDULO 3: CLASIFICACIÓN AUTOMÁTICA
# ─────────────────────────────────────────────

# Palabras clave para clasificación por tipo de estudio
CLASSIFICATION_RULES = {
    "RCT_primario": [
        r'\brandomis[ei]d\b', r'\bplacebo.controlled\b', r'\bblind(ed)?\b',
        r'\brandom(ized|ised)\s+(clinical|controlled)\s+trial\b',
        r'\bRCT\b', r'\bensayo\s+cl[ií]nico\b', r'\brandomizado\b',
        r'\bprimary\s+(result|endpoint|outcome)\b',
    ],
    "RCT_secundario": [
        r'\bsubgroup\s+anal', r'\bpost.hoc\b', r'\bsecondary\s+anal',
        r'\bsub-?study\b', r'\bpre.specified\b', r'\bpost\s+hoc\b',
    ],
    "meta-analisis": [
        r'\bmeta.anal', r'\bsystematic\s+review\b', r'\bpooled\s+anal',
        r'\bsystematic\b.*\breview\b', r'\bmetaan[aá]lisis\b',
    ],
    "registro_observacional": [
        r'\bregist(ry|er|ro)\b', r'\bcohort\b', r'\bobservational\b',
        r'\bretrospective\b', r'\bprospective\s+(cohort|observational)\b',
        r'\bepidemiolog\b',
    ],
    "guia_clinica": [
        r'\bguideline\b', r'\brecommendation\b', r'\bconsensus\s+(statement|document)\b',
        r'\bgu[ií]a\s+(cl[ií]nica|de\s+pr[aá]ctica)\b',
    ],
}

def classify_reference(record: dict) -> str:
    """
    Clasifica una referencia según palabras clave en título y texto crudo.
    Devuelve la clasificación como string.
    """
    text_to_search = " ".join([
        record.get("title", ""),
        record.get("ref_raw", ""),
        record.get("pub_type_raw", "")
    ]).lower()

    # Verificar primero RCT secundario antes que primario
    for study_type in ["RCT_secundario", "meta-analisis", "guia_clinica", 
                        "registro_observacional", "RCT_primario"]:
        patterns = CLASSIFICATION_RULES[study_type]
        for pat in patterns:
            if re.search(pat, text_to_search, re.IGNORECASE):
                return study_type

    return "otro/no_clasificado"


# ─────────────────────────────────────────────
# MÓDULO 4: EXPORTACIÓN A EXCEL
# ─────────────────────────────────────────────

COLORS = {
    "header_bg": "1F4E79",
    "header_fg": "FFFFFF",
    "rct_primary": "E2EFDA",    # verde claro
    "rct_secondary": "FFF2CC",  # amarillo claro
    "meta": "DAE8FC",           # azul claro
    "registry": "F8CECC",       # rosa claro
    "guideline": "E1D5E7",      # lila claro
    "other": "F5F5F5",          # gris claro
    "subheader": "BDD7EE",
}

STUDY_TYPE_COLORS = {
    "RCT_primario": COLORS["rct_primary"],
    "RCT_secundario": COLORS["rct_secondary"],
    "meta-analisis": COLORS["meta"],
    "registro_observacional": COLORS["registry"],
    "guia_clinica": COLORS["guideline"],
    "otro/no_clasificado": COLORS["other"],
}

def style_header(cell, bg_color=None, fg_color="FFFFFF", bold=True):
    bg = bg_color or COLORS["header_bg"]
    cell.font = Font(bold=bold, color=fg_color, name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_cell(cell, row_color=None, wrap=False):
    cell.font = Font(name="Arial", size=9)
    if row_color:
        cell.fill = PatternFill("solid", start_color=row_color)
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)

def add_thin_border(ws, row, col_start, col_end):
    thin = Side(style="thin", color="CCCCCC")
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = Border(bottom=thin)


def export_to_excel(records: list[dict], output_path: str):
    wb = Workbook()

    # ── Hoja 1: Base de datos completa ──────────────────────────────
    ws_all = wb.active
    ws_all.title = "Todas las referencias"

    columns = [
        ("N°", 5), ("Autores", 30), ("Año", 6), ("Título", 50),
        ("Revista", 25), ("PMID", 12), ("DOI", 30),
        ("URL PubMed", 35), ("URL DOI", 35),
        ("Tipo (auto)", 20), ("Tipo (manual)", 20),
        ("Notas", 25), ("Referencia original", 50),
    ]

    # Encabezado
    ws_all.row_dimensions[1].height = 30
    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws_all.cell(row=1, column=col_idx, value=col_name)
        style_header(cell)
        ws_all.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws_all.freeze_panes = "A2"

    # Datos
    for r in records:
        row_num = r["ref_number"] + 1
        study_type = r.get("study_type_auto", "otro/no_clasificado")
        row_color = STUDY_TYPE_COLORS.get(study_type, COLORS["other"])

        values = [
            r.get("ref_number", ""),
            r.get("authors", ""),
            r.get("year", ""),
            r.get("title", ""),
            r.get("journal", ""),
            r.get("pmid", ""),
            r.get("doi", ""),
            r.get("pubmed_url", ""),
            r.get("doi_url", ""),
            r.get("study_type_auto", ""),
            r.get("study_type", ""),  # campo para corrección manual
            r.get("notes", ""),
            r.get("ref_raw", ""),
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws_all.cell(row=row_num, column=col_idx, value=val)
            style_cell(cell, row_color=row_color, wrap=(col_idx in [4, 13]))
            # Hipervínculos
            if col_idx == 8 and val:
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")
            if col_idx == 9 and val:
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")

        add_thin_border(ws_all, row_num, 1, len(columns))
        ws_all.row_dimensions[row_num].height = 40

    # ── Hoja 2: Solo ECAs primarios ─────────────────────────────────
    ws_rct = wb.create_sheet("ECAs primarios")
    rct_records = [r for r in records if r.get("study_type_auto") == "RCT_primario"]
    _write_rct_sheet(ws_rct, rct_records)

    # ── Hoja 3: Resumen por tipo ────────────────────────────────────
    ws_sum = wb.create_sheet("Resumen")
    _write_summary_sheet(ws_sum, records)

    # ── Hoja 4: Instrucciones ───────────────────────────────────────
    ws_help = wb.create_sheet("Instrucciones")
    _write_instructions_sheet(ws_help)

    wb.save(output_path)
    print(f"[OK] Excel guardado: {output_path}")


def _write_rct_sheet(ws, records):
    ws.title = "ECAs primarios"
    cols = [
        ("N°", 5), ("Autores", 30), ("Año", 6), ("Título", 50),
        ("Revista", 25), ("PMID", 12), ("DOI", 30),
        ("URL PubMed", 35), ("URL DOI", 35), ("Notas", 30),
    ]
    ws.row_dimensions[1].height = 30
    for col_idx, (col_name, col_width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        style_header(cell, bg_color="375623")
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.freeze_panes = "A2"

    for row_num, r in enumerate(records, 2):
        values = [
            r.get("ref_number"), r.get("authors"), r.get("year"),
            r.get("title"), r.get("journal"), r.get("pmid"), r.get("doi"),
            r.get("pubmed_url"), r.get("doi_url"), r.get("notes"),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            style_cell(cell, row_color=COLORS["rct_primary"], wrap=(col_idx == 4))
            if col_idx == 8 and val:
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")
            if col_idx == 9 and val:
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")
        ws.row_dimensions[row_num].height = 40


def _write_summary_sheet(ws, records):
    ws.title = "Resumen"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15

    headers = ["Tipo de estudio", "N referencias", "% del total"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        style_header(cell)

    type_counts = {}
    for r in records:
        t = r.get("study_type_auto", "otro/no_clasificado")
        type_counts[t] = type_counts.get(t, 0) + 1

    total = len(records)
    for row_idx, (study_type, count) in enumerate(sorted(type_counts.items()), 2):
        pct = count / total * 100 if total else 0
        row_color = STUDY_TYPE_COLORS.get(study_type, COLORS["other"])
        ws.cell(row=row_idx, column=1, value=study_type).fill = PatternFill("solid", start_color=row_color)
        ws.cell(row=row_idx, column=2, value=count)
        ws.cell(row=row_idx, column=3, value=f"{pct:.1f}%")

    # Total
    total_row = len(type_counts) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, name="Arial")
    ws.cell(row=total_row, column=2, value=total).font = Font(bold=True, name="Arial")
    ws.cell(row=total_row, column=3, value="100%").font = Font(bold=True, name="Arial")

    # Leyenda de colores
    ws.cell(row=total_row + 2, column=1, value="LEYENDA DE COLORES:").font = Font(bold=True, name="Arial", size=9)
    for idx, (st, color) in enumerate(STUDY_TYPE_COLORS.items(), total_row + 3):
        cell = ws.cell(row=idx, column=1, value=st)
        cell.fill = PatternFill("solid", start_color=color)
        cell.font = Font(name="Arial", size=9)


def _write_instructions_sheet(ws):
    ws.title = "Instrucciones"
    ws.column_dimensions["A"].width = 80

    lines = [
        ("GUÍA DE USO DEL PIPELINE DE REFERENCIAS", True, COLORS["header_bg"], "FFFFFF"),
        ("", False, None, None),
        ("MÓDULO 1 — Extracción del PDF", True, COLORS["subheader"], "000000"),
        ("El script extrae automáticamente la sección 'References' del PDF de la guía.", False, None, None),
        ("Detecta referencias numeradas (1. Autor... o 1 Autor...) y las limpia.", False, None, None),
        ("", False, None, None),
        ("MÓDULO 2 — Enriquecimiento de metadatos", True, COLORS["subheader"], "000000"),
        ("Cada referencia se busca en CrossRef (texto libre) y PubMed (query por título+año).", False, None, None),
        ("Se extraen: PMID, DOI, título, autores, año, revista.", False, None, None),
        ("Se añaden URLs clicables a PubMed y DOI.", False, None, None),
        ("NOTA: Sin API key de NCBI el límite es 3 req/s. El script respeta este límite.", False, None, None),
        ("", False, None, None),
        ("MÓDULO 3 — Clasificación automática", True, COLORS["subheader"], "000000"),
        ("Se clasifica cada referencia por palabras clave en el título:", False, None, None),
        ("  • RCT_primario: ensayo clínico aleatorizado, publicación principal", False, None, None),
        ("  • RCT_secundario: subanálisis, post-hoc, subgrupos", False, None, None),
        ("  • meta-analisis: revisión sistemática, meta-análisis, pooled analysis", False, None, None),
        ("  • registro_observacional: registro, cohorte, observacional, retrospectivo", False, None, None),
        ("  • guia_clinica: guideline, consensus statement", False, None, None),
        ("  • otro/no_clasificado: no coincide con ningún patrón", False, None, None),
        ("La columna 'Tipo (manual)' permite correcciones manuales.", False, None, None),
        ("", False, None, None),
        ("MÓDULO 4 — Excel estructurado", True, COLORS["subheader"], "000000"),
        ("Hoja 'Todas las referencias': base de datos completa con código de colores por tipo.", False, None, None),
        ("Hoja 'ECAs primarios': solo los ECAs primarios identificados.", False, None, None),
        ("Hoja 'Resumen': tabla de frecuencias por tipo de estudio.", False, None, None),
        ("", False, None, None),
        ("USO EN LÍNEA DE COMANDOS", True, COLORS["subheader"], "000000"),
        ("  python guideline_pipeline.py guia.pdf output.xlsx", False, None, None),
        ("  python guideline_pipeline.py guia.pdf              (usa 'references_db.xlsx' por defecto)", False, None, None),
        ("", False, None, None),
        ("PARA GUÍAS CON API KEY DE NCBI (>3 req/s)", True, COLORS["subheader"], "000000"),
        ("  Añadir al entorno: export NCBI_API_KEY=tu_clave", False, None, None),
        ("  Obtener gratis en: https://www.ncbi.nlm.nih.gov/account/", False, None, None),
    ]

    for row_idx, (text, bold, bg, fg) in enumerate(lines, 1):
        cell = ws.cell(row=row_idx, column=1, value=text)
        cell.font = Font(
            bold=bold, name="Arial", size=10,
            color=fg if fg else "000000"
        )
        if bg:
            cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row_idx].height = 18


# ─────────────────────────────────────────────
# PIPELINE PRINCIPAL
# ─────────────────────────────────────────────

def run_pipeline(pdf_path: str, output_path: str = "references_db.xlsx"):
    print(f"\n{'='*60}")
    print(f"  PIPELINE DE REFERENCIAS - GUÍAS CLÍNICAS")
    print(f"{'='*60}")
    print(f"  PDF: {pdf_path}")
    print(f"  Output: {output_path}\n")

    # PASO 1: Extraer referencias
    print("[PASO 1] Extrayendo referencias del PDF...")
    raw_refs = extract_references_from_pdf(pdf_path)

    if not raw_refs:
        print("[ERROR] No se encontraron referencias. Verificar formato del PDF.")
        return

    # PASO 2 + 3: Enriquecer y clasificar
    print(f"\n[PASO 2+3] Enriqueciendo {len(raw_refs)} referencias con PubMed + CrossRef...")
    print("  (Esto puede tardar varios minutos para guías con muchas referencias)\n")

    records = []
    for idx, ref_text in enumerate(raw_refs, 1):
        print(f"  [{idx}/{len(raw_refs)}] {ref_text[:80]}...", end="\r")
        record = enrich_reference(ref_text, idx)
        record["study_type_auto"] = classify_reference(record)
        records.append(record)

    print(f"\n\n[INFO] Clasificación:")
    type_counts = {}
    for r in records:
        t = r["study_type_auto"]
        type_counts[t] = type_counts.get(t, 0) + 1
    for t, c in sorted(type_counts.items()):
        print(f"  {t}: {c}")

    # PASO 4: Exportar
    print(f"\n[PASO 4] Exportando a Excel: {output_path}")
    export_to_excel(records, output_path)

    print(f"\n{'='*60}")
    print(f"  COMPLETADO: {len(records)} referencias procesadas")
    print(f"  ECAs primarios identificados: {type_counts.get('RCT_primario', 0)}")
    print(f"{'='*60}\n")

    return records


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("USO: python guideline_pipeline.py <ruta_pdf> [output.xlsx]")
        sys.exit(1)

    pdf_file = sys.argv[1]
    out_file = sys.argv[2] if len(sys.argv) > 2 else "references_db.xlsx"
    run_pipeline(pdf_file, out_file)
